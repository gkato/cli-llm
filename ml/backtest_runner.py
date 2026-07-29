"""Run a V12 backtest against a local vLLM endpoint and write the result XLSX.

Reads a source backtest XLSX (standard format: Case name / PII data / Asserts
/ Output), sends each case to a vLLM-compatible endpoint, and writes a new
XLSX with the Output column populated by the model's responses.

Bypasses llm-playground entirely — useful when the proxy/network path is
unreliable (Cloudflare 120s timeouts on RunPod proxy URLs, SSH forwarding
disabled, etc.).

After the run, score with the same analyzer used for the other backtests:
    python3 -c "
    import openpyxl, json, re
    from collections import Counter
    ...
    "

Or use --summary which prints a quick pass-rate breakdown at the end.

Usage:
    python -m ml.backtest_runner \\
        --input         data/backtest-Merged_Prompt_V12_Qwen3_5_27B_AWQ-2026-05-12.xlsx \\
        --output        data/backtest-Merged_Prompt_V12_Qwen35_AWQ_psa-fix-2026-05-12.xlsx \\
        --short-prompt  data/prompts/V12_distill_short.txt \\
        --base-url      http://localhost:8000/v1 \\
        --model         psa-fix \\
        --summary
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.request
from collections import Counter
from pathlib import Path

import openpyxl  # type: ignore[import-not-found]


# ─── JSON extraction (same as distill_focus / distill_eval) ───────────────────

def find_last_top_level_json(text: str) -> str | None:
    last = None
    i, n = 0, len(text)
    while i < n:
        if text[i] != "{":
            i += 1
            continue
        depth, in_str, esc, end = 0, False, False, -1
        for j in range(i, n):
            c = text[j]
            if esc:
                esc = False
                continue
            if c == "\\":
                esc = True
                continue
            if c == '"':
                in_str = not in_str
                continue
            if in_str:
                continue
            if c == "{":
                depth += 1
            elif c == "}":
                depth -= 1
                if depth == 0:
                    end = j
                    break
        if end == -1:
            break
        try:
            json.loads(text[i:end + 1])
            last = text[i:end + 1]
        except (json.JSONDecodeError, ValueError):
            pass
        i = end + 1
    return last


def jget(obj, path):
    if path.startswith("$."):
        path = path[2:]
    elif path.startswith("$"):
        path = path[1:]
    cur = obj
    for part in path.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(part)
        if cur is None:
            return None
    return cur


def parse_asserts(text: str):
    out = []
    if not text:
        return out
    for chunk in text.split(" ; "):
        parts = [p.strip() for p in chunk.split("|")]
        if len(parts) < 3:
            continue
        m_path = re.search(r"@\s*(\S+)", parts[1])
        m_exp = re.search(r"expected\s*=\s*(.*)$", parts[2])
        if m_path and m_exp:
            out.append((m_path.group(1).strip(),
                        m_exp.group(1).strip().strip('"').strip("'")))
    return out


# ─── vLLM caller ──────────────────────────────────────────────────────────────

def call_vllm(base_url: str, model: str, system: str, user: str,
              max_tokens: int, timeout: int,
              enable_thinking: bool, api_key: str | None,
              temperature: float = 0.0, top_p: float | None = None,
              top_k: int | None = None) -> tuple[str, dict]:
    """POST to /v1/chat/completions. Returns (content_text, usage_dict).

    Raises on HTTP errors so the caller can record them.

    Note: Qwen3 with thinking enabled REQUIRES non-greedy sampling
    (temperature=0 traps the model in deterministic reasoning loops).
    Use temperature=0.6, top_p=0.95, top_k=20 (the model's recommended config).
    """
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
        "chat_template_kwargs": {"enable_thinking": enable_thinking},
    }
    if top_p is not None:
        body["top_p"] = top_p
    if top_k is not None:
        body["top_k"] = top_k
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    req = urllib.request.Request(
        f"{base_url}/chat/completions",
        data=json.dumps(body).encode(),
        headers=headers,
    )
    resp = urllib.request.urlopen(req, timeout=timeout)
    data = json.loads(resp.read())
    msg = data["choices"][0]["message"]
    content = msg.get("content") or ""
    # If thinking is on, content might be empty but reasoning populated — record both
    reasoning = msg.get("reasoning") or ""
    # Combine for downstream parsing (extractor will handle <think> stripping)
    full = content if content else reasoning
    usage = data.get("usage") or {}
    usage["finish_reason"] = data["choices"][0].get("finish_reason")
    return full, usage


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    p = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--input", required=True, type=Path,
                   help="Source backtest XLSX (Case name/PII data/Asserts/Output)")
    p.add_argument("--output", required=True, type=Path,
                   help="Destination XLSX — same structure, new Output column")
    p.add_argument("--short-prompt", required=True, type=Path,
                   help="System prompt text file (V12 short prompt)")
    p.add_argument("--base-url", default="http://localhost:8000/v1")
    p.add_argument("--model", required=True,
                   help="Model name registered in vLLM (e.g. 'psa-fix' for the adapter)")
    p.add_argument("--api-key", default=None,
                   help="Bearer token if vLLM was started with --api-key")
    p.add_argument("--max-tokens", type=int, default=2048,
                   help="max_tokens for each response (default 2048)")
    p.add_argument("--enable-thinking", action="store_true",
                   help="Allow Qwen3 reasoning. By default we disable it for "
                        "faster + cleaner JSON output (matches our training).")
    p.add_argument("--temperature", type=float, default=0.0,
                   help="Sampling temperature. Default 0 (greedy) — but with "
                        "--enable-thinking you MUST use 0.6 or the model gets "
                        "stuck in deterministic reasoning loops.")
    p.add_argument("--top-p", type=float, default=None,
                   help="Top-p sampling (Qwen3 recommended: 0.95 with thinking).")
    p.add_argument("--top-k", type=int, default=None,
                   help="Top-k sampling (Qwen3 recommended: 20 with thinking).")
    p.add_argument("--timeout", type=int, default=600,
                   help="Per-request timeout in seconds (default 600 = 10 min)")
    p.add_argument("--limit", type=int, default=None,
                   help="Only run first N cases (smoke test)")
    p.add_argument("--summary", action="store_true",
                   help="Print pass-rate breakdown after the run")
    args = p.parse_args()

    system = args.short_prompt.read_text().strip()
    print(f"== Backtest runner ==")
    print(f"  source:        {args.input}")
    print(f"  destination:   {args.output}")
    print(f"  endpoint:      {args.base_url}")
    print(f"  model:         {args.model}")
    print(f"  thinking:      {'ON' if args.enable_thinking else 'OFF (recommended)'}")
    print(f"  max_tokens:    {args.max_tokens}")
    print(f"  timeout:       {args.timeout}s")
    print(f"  system prompt: {len(system)} chars")
    print()

    # Read source
    src = openpyxl.load_workbook(args.input, data_only=True)
    src_ws = src["Results"]
    headers = [c.value for c in src_ws[1]]
    cases = list(src_ws.iter_rows(min_row=2, values_only=True))
    if args.limit:
        cases = cases[: args.limit]
    cases = [c for c in cases if c[0] is not None]
    print(f"Loaded {len(cases)} cases")
    print()

    # Prep destination
    dst = openpyxl.Workbook()
    dst_ws = dst.active
    dst_ws.title = "Results"
    dst_ws.append(headers)
    args.output.parent.mkdir(parents=True, exist_ok=True)

    # Stats
    n_ok = n_err = 0
    pass_count = 0
    field_correct: Counter[str] = Counter()
    field_total: Counter[str] = Counter()
    pred_class: list[str] = []
    gold_class: list[str] = []
    t_run_start = time.monotonic()

    for i, (name, pii, asserts, _old) in enumerate(cases, 1):
        t0 = time.monotonic()
        try:
            output_text, usage = call_vllm(
                args.base_url, args.model, system, pii or "",
                max_tokens=args.max_tokens, timeout=args.timeout,
                enable_thinking=args.enable_thinking, api_key=args.api_key,
                temperature=args.temperature, top_p=args.top_p, top_k=args.top_k,
            )
            n_ok += 1
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            output_text = f"ERROR: {type(e).__name__}: {e}"
            usage = {"prompt_tokens": 0, "completion_tokens": 0, "finish_reason": "error"}
            n_err += 1
        elapsed = time.monotonic() - t0

        # Write output row immediately so we can resume after a crash
        dst_ws.append([name, pii, asserts, output_text])
        dst.save(args.output)

        # On-the-fly scoring
        pred_obj = None
        if output_text and not output_text.startswith("ERROR:"):
            block = find_last_top_level_json(output_text)
            if block:
                try:
                    pred_obj = json.loads(block)
                except json.JSONDecodeError:
                    pred_obj = None

        case_passed = pred_obj is not None
        assert_results = []
        for path, expected in parse_asserts(asserts or ""):
            actual = jget(pred_obj or {}, path)
            actual_s = "" if actual is None else str(actual)
            match = actual_s.strip() == expected.strip()
            field = path.split(".")[-1]
            field_total[field] += 1
            if match:
                field_correct[field] += 1
            else:
                case_passed = False
            if "document_classification" in path:
                gold_class.append(expected)
                pred_class.append(actual_s)
            assert_results.append((field, match, expected, actual_s))

        if case_passed and pred_obj is not None:
            pass_count += 1
            mark = "✓"
        elif pred_obj is None:
            mark = "?"  # couldn't parse
        else:
            mark = "✗"

        usage_str = f"in={usage.get('prompt_tokens',0):>5} out={usage.get('completion_tokens',0):>4}"
        running_pass = pass_count / i if i else 0
        print(f"  [{i:>2}/{len(cases)}] [{mark}] {elapsed:>5.1f}s  {usage_str}  "
              f"running={100*running_pass:.1f}%  {str(name)[:50]}")

    total = time.monotonic() - t_run_start
    print()
    print(f"== Run complete in {total/60:.1f} min ==")
    print(f"  Output XLSX: {args.output}")
    print(f"  OK:          {n_ok}")
    print(f"  Errors:      {n_err}")
    print(f"  Cases passed: {pass_count}/{len(cases)} ({100*pass_count/len(cases):.1f}%)")

    if args.summary:
        print()
        print("== Per-field accuracy ==")
        for field in sorted(field_total.keys(), key=lambda f: -field_total[f]):
            c, t = field_correct[field], field_total[field]
            print(f"  {field:<40}  {c:>3}/{t:<3}  ({100*c/t:.1f}%)")

        print()
        print("== document_classification confusion ==")
        classes = sorted(set(gold_class) | set(pred_class))
        if classes:
            head = f"  {'gold↓ / pred→':<30}" + "".join(f"  {c[:15]:>17}" for c in classes)
            print(head)
            for g in classes:
                row = f"  {g[:30]:<30}"
                for pp in classes:
                    n = sum(1 for gc, pc in zip(gold_class, pred_class)
                            if gc == g and pc == pp)
                    row += f"  {n:>17}"
                print(row)


if __name__ == "__main__":
    main()
